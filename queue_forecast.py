#!/usr/bin/env python3
"""
Queue Forecasting System
========================
Implements Holt-Winters exponential smoothing with weekly seasonality
for ticket demand forecasting, following the Queue Projections methodology.

Usage:
    from queue_forecast import QueueForecaster
    
    forecaster = QueueForecaster()
    forecaster.load_data('daily_table.csv')
    forecaster.generate_forecast(horizon=30)
    forecaster.export_excel('output.xlsx')
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional


@dataclass
class ForecastConfig:
    """Configuration parameters for the forecasting model."""
    alpha: float = 0.2       # Level smoothing
    beta: float = 0.1        # Trend smoothing
    gamma: float = 0.2       # Seasonal smoothing
    seasonality: int = 7     # Weekly seasonality
    shares_window: int = 28  # Days for share calculation
    bb_lambda: float = 0.1   # Beta-binomial decay
    horizon: int = 30        # Forecast horizon


class HoltWinters:
    """Holt-Winters additive seasonal forecasting."""
    
    def __init__(self, alpha=0.2, beta=0.1, gamma=0.2, m=7):
        self.alpha = alpha
        self.beta = beta
        self.gamma = gamma
        self.m = m
    
    def fit_forecast(self, y: np.ndarray, horizon: int) -> List[float]:
        n = len(y)
        if n < 2 * self.m:
            return [max(0, np.mean(y)) for _ in range(horizon)]
        
        # Initialize
        l = [np.mean(y[:self.m])]
        b = [(np.mean(y[self.m:2*self.m]) - np.mean(y[:self.m])) / self.m]
        s = list(y[:self.m] - l[0])
        
        # Fit
        for t in range(1, n):
            s_tm = s[t-1] if t <= self.m else s[t-self.m]
            l_t = self.alpha * (y[t] - s_tm) + (1 - self.alpha) * (l[t-1] + b[t-1])
            l.append(l_t)
            b_t = self.beta * (l[t] - l[t-1]) + (1 - self.beta) * b[t-1]
            b.append(b_t)
            s_t = self.gamma * (y[t] - l[t]) + (1 - self.gamma) * s_tm
            s.append(s_t)
        
        # Forecast
        forecasts = []
        for h in range(1, horizon + 1):
            idx = len(s) - self.m + ((h-1) % self.m)
            s_f = s[max(0, min(idx, len(s)-1))]
            y_hat = l[-1] + h * b[-1] + s_f
            forecasts.append(max(0, y_hat))
        
        return forecasts


class BetaBinomialSmoother:
    """Beta-binomial smoothing for rate estimation."""
    
    def __init__(self, lambda_decay: float = 0.1):
        self.lambda_decay = lambda_decay
        self.alpha = 1.0
        self.beta = 1.0
    
    def update(self, successes: int, trials: int):
        self.alpha = self.lambda_decay * successes + (1 - self.lambda_decay) * self.alpha
        self.beta = self.lambda_decay * max(0, trials - successes) + (1 - self.lambda_decay) * self.beta
    
    @property
    def rate(self) -> float:
        return self.alpha / (self.alpha + self.beta)


class QueueForecaster:
    """Main forecasting class for ticket queue projections."""
    
    def __init__(self, config: Optional[ForecastConfig] = None):
        self.config = config or ForecastConfig()
        self.daily_df: Optional[pd.DataFrame] = None
        self.forecast_df: Optional[pd.DataFrame] = None
        self.vertical_shares: Dict[str, float] = {}
        self.issue_shares: Dict[str, float] = {}
        self.escalation_rate: float = 0.0
    
    def load_data(self, filepath: str):
        """Load daily table from CSV."""
        self.daily_df = pd.read_csv(filepath, parse_dates=['date'])
        return self
    
    def load_dataframe(self, df: pd.DataFrame):
        """Load from existing DataFrame."""
        self.daily_df = df.copy()
        if 'date' in self.daily_df.columns:
            self.daily_df['date'] = pd.to_datetime(self.daily_df['date'])
        return self
    
    def set_shares(self, vertical_shares: Dict[str, float], issue_shares: Dict[str, float]):
        """Set vertical and issue shares for allocation."""
        self.vertical_shares = vertical_shares
        self.issue_shares = issue_shares
        return self
    
    def compute_escalation_rate(self) -> float:
        """Compute smoothed escalation rate using Beta-Binomial."""
        smoother = BetaBinomialSmoother(self.config.bb_lambda)
        for _, row in self.daily_df.iterrows():
            smoother.update(int(row['E_t']), int(row['A_t']))
        self.escalation_rate = smoother.rate
        return self.escalation_rate
    
    def generate_forecast(self, horizon: Optional[int] = None) -> pd.DataFrame:
        """Generate full forecast."""
        if self.daily_df is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        h = horizon or self.config.horizon
        hw = HoltWinters(
            self.config.alpha, self.config.beta,
            self.config.gamma, self.config.seasonality
        )
        
        # Forecast arrivals and closures
        arr_forecast = hw.fit_forecast(self.daily_df['A_t'].values, h)
        cls_forecast = hw.fit_forecast(self.daily_df['C_t'].values, h)
        
        # Escalation rate
        self.compute_escalation_rate()
        esc_forecast = [a * self.escalation_rate for a in arr_forecast]
        
        # Backlog projection
        current_backlog = self.daily_df['B_t'].iloc[-1]
        backlog = []
        running = current_backlog
        for i in range(h):
            running = max(0, running + arr_forecast[i] - cls_forecast[i])
            backlog.append(running)
        
        # Create forecast DataFrame
        last_date = self.daily_df['date'].max()
        dates = pd.date_range(start=last_date + timedelta(days=1), periods=h, freq='D')
        
        self.forecast_df = pd.DataFrame({
            'date': dates,
            'A_hat': arr_forecast,
            'C_hat': cls_forecast,
            'E_rate': self.escalation_rate,
            'E_hat': esc_forecast,
            'B_hat': backlog
        })
        
        return self.forecast_df
    
    def allocate_by_vertical(self) -> pd.DataFrame:
        """Allocate arrivals forecast by vertical."""
        if self.forecast_df is None:
            raise ValueError("Generate forecast first.")
        
        result = self.forecast_df[['date', 'A_hat']].copy()
        for vert, share in self.vertical_shares.items():
            result[vert] = result['A_hat'] * share
        return result
    
    def allocate_by_issue(self) -> pd.DataFrame:
        """Allocate arrivals forecast by issue."""
        if self.forecast_df is None:
            raise ValueError("Generate forecast first.")
        
        result = self.forecast_df[['date', 'A_hat']].copy()
        for issue, share in self.issue_shares.items():
            result[issue] = result['A_hat'] * share
        return result
    
    def summary(self) -> Dict:
        """Return forecast summary metrics."""
        if self.forecast_df is None:
            raise ValueError("Generate forecast first.")
        
        current_backlog = self.daily_df['B_t'].iloc[-1]
        return {
            'training_start': self.daily_df['date'].min(),
            'training_end': self.daily_df['date'].max(),
            'training_days': len(self.daily_df),
            'forecast_start': self.forecast_df['date'].min(),
            'forecast_end': self.forecast_df['date'].max(),
            'forecast_days': len(self.forecast_df),
            'total_arrivals': self.forecast_df['A_hat'].sum(),
            'total_closures': self.forecast_df['C_hat'].sum(),
            'total_escalations': self.forecast_df['E_hat'].sum(),
            'escalation_rate': self.escalation_rate,
            'current_backlog': current_backlog,
            'final_backlog': self.forecast_df['B_hat'].iloc[-1],
            'backlog_change': self.forecast_df['B_hat'].iloc[-1] - current_backlog
        }
    
    def export_excel(self, filepath: str):
        """Export forecast to Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2F5496')
        
        # Queue Forecast sheet
        ws = wb.active
        ws.title = "Queue Forecast"
        headers = ['Date', 'A_hat', 'E_rate', 'E_hat', 'C_hat', 'B_hat']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        
        for i, row in self.forecast_df.iterrows():
            ws.cell(row=i+2, column=1, value=row['date'].date())
            ws.cell(row=i+2, column=2, value=round(row['A_hat'], 1))
            ws.cell(row=i+2, column=3, value=round(row['E_rate'], 4))
            ws.cell(row=i+2, column=4, value=round(row['E_hat'], 1))
            ws.cell(row=i+2, column=5, value=round(row['C_hat'], 1))
            ws.cell(row=i+2, column=6, value=round(row['B_hat'], 0))
        
        # By Vertical sheet (if shares available)
        if self.vertical_shares:
            ws2 = wb.create_sheet("By Vertical")
            vertical_df = self.allocate_by_vertical()
            for col, h in enumerate(vertical_df.columns, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
            for i, row in vertical_df.iterrows():
                for col, val in enumerate(row, 1):
                    if col == 1:
                        ws2.cell(row=i+2, column=col, value=val.date() if hasattr(val, 'date') else val)
                    else:
                        ws2.cell(row=i+2, column=col, value=round(val, 1) if isinstance(val, float) else val)
        
        wb.save(filepath)
        return filepath


def build_daily_table(ticket_df: pd.DataFrame, 
                      date_col: str = 'date',
                      ticket_col: str = 'ticket_number',
                      status_col: str = 'status',
                      account_col: str = 'account_type',
                      issue_col: str = 'issue_tag') -> pd.DataFrame:
    """
    Build daily table from raw ticket data.
    
    Parameters:
    - ticket_df: DataFrame with ticket records
    - date_col: Column name for dates
    - ticket_col: Column name for ticket IDs
    - status_col: Column name for status
    - account_col: Column name for account type
    - issue_col: Column name for issue tag
    
    Returns:
    - DataFrame with columns: date, A_t, C_t, E_t, B_t
    """
    df = ticket_df.copy()
    df[date_col] = pd.to_datetime(df[date_col])
    
    # Normalize status
    def norm_status(s):
        if pd.isna(s):
            return 'Unknown'
        s = str(s).upper().strip()
        if 'CLOSED' in s:
            return 'Closed'
        elif 'ESCALAT' in s:
            return 'Escalated'
        elif 'WFC' in s:
            return 'WFC'
        elif 'REASSIGN' in s:
            return 'Reassigned'
        return s
    
    df[status_col] = df[status_col].apply(norm_status)
    
    # First occurrence dates
    arrivals = df.groupby(ticket_col)[date_col].min().reset_index()
    arrivals.columns = [ticket_col, 'arrival_date']
    
    closed = df[df[status_col] == 'Closed'].groupby(ticket_col)[date_col].min().reset_index()
    closed.columns = [ticket_col, 'close_date']
    
    escalated = df[df[status_col] == 'Escalated'].groupby(ticket_col)[date_col].min().reset_index()
    escalated.columns = [ticket_col, 'escalation_date']
    
    tickets = arrivals.merge(closed, on=ticket_col, how='left')
    tickets = tickets.merge(escalated, on=ticket_col, how='left')
    
    # Build daily table
    date_range = pd.date_range(df[date_col].min(), df[date_col].max(), freq='D')
    daily = []
    backlog = 0
    
    for dt in date_range:
        A_t = len(tickets[tickets['arrival_date'] == dt])
        C_t = len(tickets[tickets['close_date'] == dt])
        E_t = len(tickets[tickets['escalation_date'] == dt])
        backlog = max(0, backlog + A_t - C_t)
        daily.append({'date': dt, 'A_t': A_t, 'C_t': C_t, 'E_t': E_t, 'B_t': backlog})
    
    return pd.DataFrame(daily)


if __name__ == '__main__':
    # Example usage
    print("Queue Forecasting Module")
    print("Import and use: from queue_forecast import QueueForecaster")
